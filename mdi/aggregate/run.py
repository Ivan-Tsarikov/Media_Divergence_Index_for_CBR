from __future__ import annotations

import argparse
from pathlib import Path
import pandas as pd
import numpy as np

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .metrics import majority_stance, avg_strength, top_reasons, compute_divergence


# ---------- I/O ----------

def read_events_optional(path: str | None) -> pd.DataFrame | None:
    """
    Optional events table to enrich output (dates, decision, new_rate, links).
    Required columns (if provided): event_id (or event_date_time to derive), optionally decision/new_rate/urls.
    """
    if not path:
        return None

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"events file not found: {path}")

    if p.suffix.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(p)
    else:
        df = pd.read_csv(p, encoding="utf-8-sig")

    if "event_id" not in df.columns:
        if "event_date_time" in df.columns:
            dt = pd.to_datetime(df["event_date_time"], errors="coerce")
            df["event_id"] = dt.dt.strftime("cbr_%Y%m%d")
        else:
            raise ValueError("events file must contain 'event_id' or 'event_date_time'")

    if "event_date_time" in df.columns:
        df["event_date_time"] = pd.to_datetime(df["event_date_time"], errors="coerce")

    # Keep only useful columns (if exist)
    keep = [c for c in [
        "event_id", "event_date_time", "decision", "new_rate", "cbr_press_url", "cbr_resume_url"
    ] if c in df.columns]
    return df[keep].copy()


def read_annotations(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, encoding="utf-8-sig")

    # Normalize core columns
    df["source_type"] = df["source_type"].astype(str).str.lower()
    df["stance"] = df["stance"].astype(str).str.lower()
    df["status"] = df["status"].astype(str).str.lower()

    if "mentions_key_rate" in df.columns:
        df["mentions_key_rate"] = df["mentions_key_rate"].fillna(False).astype(bool)
    else:
        df["mentions_key_rate"] = True

    # Some columns may not exist; don't crash
    for col in ["published_at", "title", "source_name", "evidence", "reasons", "strength", "doc_id", "event_id"]:
        if col not in df.columns:
            df[col] = np.nan

    return df


# ---------- Excel formatting helpers ----------

def _set_auto_filter_and_freeze(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _auto_fit_columns(ws, max_width: int = 70):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        width = min(max_len + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, width)


def _wrap_columns(ws, col_names: list[str]):
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    name_to_idx = {name: i for i, name in enumerate(header, start=1)}
    for name in col_names:
        idx = name_to_idx.get(name)
        if not idx:
            continue
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_xlsx(mdi_df: pd.DataFrame, media_qc_df: pd.DataFrame, out_path: str):
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        mdi_df.to_excel(writer, sheet_name="mdi_by_event", index=False)
        media_qc_df.to_excel(writer, sheet_name="media_qc_by_doc", index=False)

        ws1 = writer.sheets["mdi_by_event"]
        ws2 = writer.sheets["media_qc_by_doc"]

        _set_auto_filter_and_freeze(ws1)
        _set_auto_filter_and_freeze(ws2)

        # Wrap long text fields
        _wrap_columns(ws1, ["official_top_reasons", "media_top_reasons", "cbr_press_url", "cbr_resume_url"])
        _wrap_columns(ws2, ["title", "evidence"])

        _auto_fit_columns(ws1, max_width=55)
        _auto_fit_columns(ws2, max_width=70)


# ---------- Aggregation ----------

def build_media_qc(media: pd.DataFrame) -> pd.DataFrame:
    m = media.copy()
    m["ok"] = m["status"].eq("ok")
    m["usable"] = m["ok"] & m["mentions_key_rate"] & ~m["stance"].isin(["irrelevant"])

    m["excluded_reason"] = ""
    m.loc[~m["ok"], "excluded_reason"] = "status_not_ok"
    m.loc[m["ok"] & ~m["mentions_key_rate"], "excluded_reason"] = "mentions_key_rate_false"
    m.loc[m["ok"] & m["mentions_key_rate"] & m["stance"].isin(["irrelevant"]), "excluded_reason"] = "stance_irrelevant"
    m.loc[m["usable"], "excluded_reason"] = ""

    # Keep only informative columns for human audit
    cols = [
        "event_id", "doc_id", "published_at", "source_name", "title",
        "stance", "strength", "mentions_key_rate", "status", "excluded_reason", "evidence"
    ]
    cols = [c for c in cols if c in m.columns]
    out = m[cols].copy()

    # Sort for readability
    return out.sort_values(["event_id", "source_name", "published_at", "doc_id"], na_position="last")


def aggregate(events_path: str | None, annotations_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    events = read_events_optional(events_path)
    ann = read_annotations(annotations_path)

    # Split
    cbr = ann[ann["source_type"] == "cbr"].copy()
    media = ann[ann["source_type"] == "media"].copy()

    # Flags
    cbr["ok"] = cbr["status"].eq("ok")
    cbr["usable"] = cbr["ok"] & cbr["mentions_key_rate"] & ~cbr["stance"].isin(["irrelevant"])

    media["ok"] = media["status"].eq("ok")
    media["usable"] = media["ok"] & media["mentions_key_rate"] & ~media["stance"].isin(["irrelevant"])

    # Official aggregation (usually 1 doc per event)
    official_rows = []
    for event_id, g in cbr.groupby("event_id"):
        g_used = g[g["usable"]]

        stance = g_used["stance"].iloc[0] if len(g_used) else "na"
        strength = float(pd.to_numeric(g_used["strength"], errors="coerce").iloc[0]) if len(g_used) else np.nan

        official_rows.append({
            "event_id": event_id,
            "n_cbr_used": int(len(g_used)),
            "official_stance": stance,
            "official_strength": strength,
            "official_top_reasons": top_reasons(g_used.get("reasons", pd.Series(dtype=str)), k=2),
        })
    official = pd.DataFrame(official_rows)

    # Media aggregation
    media_rows = []
    for event_id, g in media.groupby("event_id"):
        g_ok = g[g["ok"]]
        g_used = g[g["usable"]]

        media_rows.append({
            "event_id": event_id,
            "n_media_total": int(len(g)),
            "n_media_ok": int(len(g_ok)),
            "n_media_used": int(len(g_used)),
            "media_relevance_rate": float(g_ok["mentions_key_rate"].mean()) if len(g_ok) else np.nan,
            "media_stance": majority_stance(g_used["stance"]),
            "media_strength_avg": avg_strength(g_used["strength"]),
            "media_top_reasons": top_reasons(g_used.get("reasons", pd.Series(dtype=str)), k=2),
        })
    media_agg = pd.DataFrame(media_rows)

    # Base event table for output:
    if events is None:
        # If no events table, build minimal event table from annotations
        base = pd.DataFrame({"event_id": sorted(ann["event_id"].dropna().astype(str).unique().tolist())})
    else:
        base = events.copy()

    mdi = (
        base.merge(official, on="event_id", how="left")
            .merge(media_agg, on="event_id", how="left")
    )

    mdi["divergence_stance"] = mdi.apply(
        lambda r: compute_divergence(r.get("official_stance"), r.get("media_stance")),
        axis=1
    )

    # Keep ONLY useful columns for quick evaluation
    mdi_cols = [
        "event_id",
        "event_date_time",
        "decision",
        "new_rate",
        "official_stance",
        "official_strength",
        "media_stance",
        "media_strength_avg",
        "divergence_stance",
        "n_media_used",
        "n_media_ok",
        "n_media_total",
        "media_relevance_rate",
        "official_top_reasons",
        "media_top_reasons",
        "cbr_press_url",
        "cbr_resume_url",
    ]
    mdi_cols = [c for c in mdi_cols if c in mdi.columns]
    mdi = mdi[mdi_cols].copy()

    if "event_date_time" in mdi.columns:
        mdi = mdi.sort_values("event_date_time")

    media_qc = build_media_qc(media)
    return mdi, media_qc


# ---------- CLI ----------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ann", required=True, help="Path to annotations csv (e.g., data/out/annotations_v2.csv)")
    ap.add_argument("--events", default=None, help="Optional events file (csv/xlsx) to add date/decision/rate/urls")
    ap.add_argument("--out_xlsx", default="data/out/mdi_outputs.xlsx", help="Output Excel path (2 sheets)")
    args = ap.parse_args()

    mdi_df, media_qc_df = aggregate(args.events, args.ann)
    write_xlsx(mdi_df, media_qc_df, args.out_xlsx)

    print(f"Saved Excel: {args.out_xlsx}")
    print("Sheets: mdi_by_event, media_qc_by_doc")


if __name__ == "__main__":
    main()
