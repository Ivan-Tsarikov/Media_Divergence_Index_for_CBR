from __future__ import annotations

import argparse
import csv
from pathlib import Path
import pandas as pd
import numpy as np

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


# -----------------------------
# Robust readers
# -----------------------------

def read_annotations(path: str) -> pd.DataFrame:
    p = Path(path)
    if p.suffix.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(p)
    else:
        # tolerate broken CSV quoting (just in case)
        try:
            df = pd.read_csv(p, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(
                p,
                encoding="utf-8-sig",
                engine="python",
                sep=",",
                quoting=csv.QUOTE_MINIMAL,
                on_bad_lines="skip",
            )

    # normalize expected columns
    for col in ["event_id", "doc_id", "source_type", "source_name", "published_at", "title",
                "stance", "strength", "mentions_key_rate", "status", "evidence", "reasons"]:
        if col not in df.columns:
            df[col] = np.nan

    df["event_id"] = df["event_id"].astype(str)
    df["doc_id"] = df["doc_id"].astype(str)
    df["source_type"] = df["source_type"].astype(str).str.lower()
    df["source_name"] = df["source_name"].astype(str)
    df["stance"] = df["stance"].astype(str).str.lower()
    df["status"] = df["status"].astype(str).str.lower()

    df["published_at"] = pd.to_datetime(df["published_at"], errors="coerce")
    df["mentions_key_rate"] = df["mentions_key_rate"].fillna(False).astype(bool)
    df["strength"] = pd.to_numeric(df["strength"], errors="coerce")

    # Clean strings
    df["title"] = df["title"].fillna("").astype(str).str.strip()
    df["evidence"] = df["evidence"].fillna("").astype(str).str.strip()
    df["reasons"] = df["reasons"].fillna("").astype(str).str.strip()

    return df


def read_mdi_xlsx(path: str, sheet: str = "mdi_by_event") -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"MDI xlsx not found: {path}")
    df = pd.read_excel(p, sheet_name=sheet)

    # normalize expected columns
    for col in ["event_id", "divergence_stance", "official_stance", "media_stance",
                "event_date_time", "decision", "new_rate", "n_media_used",
                "media_relevance_rate", "official_top_reasons", "media_top_reasons"]:
        if col not in df.columns:
            df[col] = np.nan

    df["event_id"] = df["event_id"].astype(str)
    df["event_date_time"] = pd.to_datetime(df["event_date_time"], errors="coerce")
    return df


# -----------------------------
# Selection logic (which events become cases)
# -----------------------------

def select_events(mdi: pd.DataFrame, max_cases: int) -> list[str]:
    """
    Priority:
      1) divergence_stance == 1
      2) media_stance == 'mixed'
      3) low coverage (n_media_used < 4)
      4) fallback: first events by date
    """
    df = mdi.copy()

    # normalize
    df["divergence_stance_num"] = pd.to_numeric(df["divergence_stance"], errors="coerce")
    df["n_media_used_num"] = pd.to_numeric(df["n_media_used"], errors="coerce")

    picked: list[str] = []

    def take(mask):
        nonlocal picked
        cand = df.loc[mask, "event_id"].dropna().astype(str).tolist()
        for e in cand:
            if e not in picked:
                picked.append(e)
            if len(picked) >= max_cases:
                break

    take(df["divergence_stance_num"].eq(1))
    if len(picked) < max_cases:
        take(df["media_stance"].astype(str).str.lower().eq("mixed"))
    if len(picked) < max_cases:
        take(df["n_media_used_num"].lt(4))
    if len(picked) < max_cases:
        # fallback by date
        df2 = df.sort_values("event_date_time")
        take(df2["event_id"].notna())

    return picked[:max_cases]


# -----------------------------
# Case assembly
# -----------------------------

def build_media_flags(media: pd.DataFrame) -> pd.DataFrame:
    m = media.copy()
    m["ok"] = m["status"].eq("ok")
    m["usable"] = m["ok"] & m["mentions_key_rate"] & ~m["stance"].isin(["irrelevant"])

    m["excluded_reason"] = ""
    m.loc[~m["ok"], "excluded_reason"] = "status_not_ok"
    m.loc[m["ok"] & ~m["mentions_key_rate"], "excluded_reason"] = "mentions_key_rate_false"
    m.loc[m["ok"] & m["mentions_key_rate"] & m["stance"].isin(["irrelevant"]), "excluded_reason"] = "stance_irrelevant"
    m.loc[m["usable"], "excluded_reason"] = ""

    m["is_used_in_media_signal"] = m["usable"].astype(bool)
    return m


def short_evidence(evidence: str, max_items: int = 2) -> str:
    if not evidence:
        return ""
    # evidence in your pipeline is "a ||| b ||| c"
    parts = [p.strip() for p in evidence.split("|||") if p.strip()]
    return " ||| ".join(parts[:max_items])


def build_outputs(mdi: pd.DataFrame, ann: pd.DataFrame, selected_events: list[str]) -> tuple[pd.DataFrame, pd.DataFrame, str]:
    # docs split
    cbr = ann[ann["source_type"] == "cbr"].copy()
    media = ann[ann["source_type"] == "media"].copy()
    media = build_media_flags(media)

    # ---- cases summary (one row per event)
    cases = mdi[mdi["event_id"].isin(selected_events)].copy()

    # readable takeaway (auto)
    def takeaway(r):
        o = str(r.get("official_stance", "")).lower()
        m = str(r.get("media_stance", "")).lower()
        d = r.get("divergence_stance", np.nan)
        if pd.isna(d):
            return "Недостаточно данных для оценки расхождения"
        if float(d) == 0:
            return f"Согласие: ЦБ={o}, СМИ={m}"
        if float(d) == 1:
            return f"Расхождение: ЦБ={o}, СМИ={m}"
        return ""

    cases["takeaway"] = cases.apply(takeaway, axis=1)

    # keep only useful columns
    keep_cases = [
        "event_id", "event_date_time", "decision", "new_rate",
        "official_stance", "media_stance", "divergence_stance",
        "n_media_used", "media_relevance_rate",
        "official_top_reasons", "media_top_reasons",
        "takeaway"
    ]
    keep_cases = [c for c in keep_cases if c in cases.columns]
    cases = cases[keep_cases].copy()

    # ---- docs table (all docs for selected events)
    docs = ann[ann["event_id"].isin(selected_events)].copy()

    # add media flags to docs
    docs = docs.merge(
        media[["doc_id", "excluded_reason", "is_used_in_media_signal"]],
        on="doc_id",
        how="left"
    )
    docs["excluded_reason"] = docs["excluded_reason"].fillna("")
    docs["is_used_in_media_signal"] = docs["is_used_in_media_signal"].fillna(False).astype(bool)

    # shorten evidence to make xlsx readable
    docs["evidence_short"] = docs["evidence"].apply(lambda x: short_evidence(str(x), max_items=2))

    keep_docs = [
        "event_id", "doc_id", "source_type", "source_name", "published_at",
        "title", "stance", "strength", "mentions_key_rate", "status",
        "is_used_in_media_signal", "excluded_reason",
        "evidence_short"
    ]
    keep_docs = [c for c in keep_docs if c in docs.columns]
    docs = docs[keep_docs].copy()
    docs = docs.sort_values(["event_id", "source_type", "source_name", "published_at"], na_position="last")

    # ---- markdown report
    md_lines = []
    md_lines.append("# Шаг 5 — кейсы Media Divergence Index\n")
    md_lines.append(f"Выбрано кейсов: {len(selected_events)}\n")

    for _, r in cases.sort_values("event_date_time").iterrows():
        eid = r["event_id"]
        md_lines.append(f"\n## {eid}\n")
        if "event_date_time" in cases.columns and pd.notna(r.get("event_date_time", None)):
            md_lines.append(f"- Дата: {pd.to_datetime(r['event_date_time']).date()}\n")
        if "decision" in cases.columns:
            md_lines.append(f"- Решение: {r.get('decision','')}\n")
        if "new_rate" in cases.columns:
            md_lines.append(f"- Новая ставка: {r.get('new_rate','')}\n")

        md_lines.append(f"- ЦБ stance: **{r.get('official_stance','')}**\n")
        md_lines.append(f"- СМИ stance: **{r.get('media_stance','')}**\n")
        md_lines.append(f"- Divergence: **{r.get('divergence_stance','')}**\n")
        md_lines.append(f"- Покрытие СМИ (used): {r.get('n_media_used','')}, relevance_rate: {r.get('media_relevance_rate','')}\n")
        md_lines.append(f"- Итог: {r.get('takeaway','')}\n")

        # add small doc listing
        sub = docs[docs["event_id"] == eid].copy()
        md_lines.append("\n### Документы\n")
        for _, d in sub.iterrows():
            src = d.get("source_type", "")
            name = d.get("source_name", "")
            title = d.get("title", "")
            stance = d.get("stance", "")
            used = d.get("is_used_in_media_signal", False)
            excl = d.get("excluded_reason", "")
            ev = d.get("evidence_short", "")

            flag = "USED" if used else ("EXCL:" + excl if excl else "—")
            md_lines.append(f"- [{src}/{name}] **{stance}** ({flag}) — {title}\n")
            if ev:
                md_lines.append(f"  - evidence: {ev}\n")

    md_report = "".join(md_lines)
    return cases, docs, md_report


# -----------------------------
# Excel writer
# -----------------------------

def _set_auto_filter_and_freeze(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _auto_fit_columns(ws, max_width: int = 70):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        width = min(max_len + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, width)


def _wrap_columns(ws, col_names: list[str]):
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    name_to_idx = {name: i for i, name in enumerate(header, start=1)}
    for name in col_names:
        idx = name_to_idx.get(name)
        if not idx:
            continue
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_xlsx(cases: pd.DataFrame, docs: pd.DataFrame, out_path: str):
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        cases.to_excel(writer, sheet_name="cases", index=False)
        docs.to_excel(writer, sheet_name="docs", index=False)

        ws1 = writer.sheets["cases"]
        ws2 = writer.sheets["docs"]
        _set_auto_filter_and_freeze(ws1)
        _set_auto_filter_and_freeze(ws2)

        _wrap_columns(ws1, ["official_top_reasons", "media_top_reasons", "takeaway"])
        _wrap_columns(ws2, ["title", "evidence_short"])

        _auto_fit_columns(ws1, max_width=55)
        _auto_fit_columns(ws2, max_width=80)


# -----------------------------
# CLI
# -----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ann", required=True, help="Path to annotations csv (e.g., data/out/annotations.csv)")
    ap.add_argument("--mdi_xlsx", required=True, help="Path to mdi_outputs.xlsx from step 4 (sheet mdi_by_event)")
    ap.add_argument("--max_cases", type=int, default=5, help="Max number of events to include as cases")
    ap.add_argument("--out_xlsx", default="data/out/cases_outputs.xlsx", help="Output Excel path (cases + docs)")
    ap.add_argument("--out_md", default="data/out/cases_report.md", help="Output Markdown report path")
    args = ap.parse_args()

    ann = read_annotations(args.ann)
    mdi = read_mdi_xlsx(args.mdi_xlsx, sheet="mdi_by_event")

    selected = select_events(mdi, max_cases=args.max_cases)
    cases_df, docs_df, md_report = build_outputs(mdi, ann, selected)

    write_xlsx(cases_df, docs_df, args.out_xlsx)
    Path(args.out_md).parent.mkdir(parents=True, exist_ok=True)
    Path(args.out_md).write_text(md_report, encoding="utf-8")

    print(f"Saved: {args.out_xlsx} (sheets: cases, docs)")
    print(f"Saved: {args.out_md}")
    print(f"Selected events: {', '.join(selected)}")


if __name__ == "__main__":
    main()
